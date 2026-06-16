"""Private workbook analysis for the ten-year portfolio workflow.

The parser intentionally summarizes workbook structure and methodology signals
without returning raw holdings, account values, notes, or ticker lists.
"""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO, StringIO
import csv
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


MAX_UPLOAD_BYTES = 15 * 1024 * 1024
MAX_SCAN_ROWS = 250
MAX_SCAN_COLUMNS = 80

PRIVATE_FIELD_PATTERNS = (
    "account", "acct", "broker", "custodian", "shares", "share", "quantity",
    "qty", "position", "market value", "value", "cost", "basis", "gain",
    "loss", "p/l", "pl", "notes", "comment", "memo", "tax lot", "lot",
    "purchase", "bought", "price paid", "owner",
)

METHODOLOGY_PATTERNS: dict[str, tuple[str, ...]] = {
    "long_term_chart": ("10 year", "10yr", "ten year", "5 year", "5yr", "chart", "trend", "slope"),
    "relative_strength": ("qqq", "nasdaq", "spy", "relative", "benchmark", "vs"),
    "momentum": ("momentum", "rsi", "moving average", "ma50", "ma200", "200 day", "50 day"),
    "quality": ("roe", "roic", "margin", "cash flow", "debt", "quality"),
    "valuation": ("pe", "p/e", "peg", "multiple", "valuation", "sales", "revenue"),
    "income": ("dividend", "yield", "payout"),
    "risk": ("drawdown", "beta", "volatility", "risk", "downside"),
    "ranking": ("rank", "score", "weight", "rating", "grade"),
    "portfolio_rules": ("target", "allocation", "rebalance", "hold", "sell", "buy"),
}

FORMULA_FUNCTION_RE = re.compile(r"\b([A-Z][A-Z0-9_.]*)\s*\(")
TICKER_LIKE_RE = re.compile(r"^[A-Z]{1,5}(?:[-.][A-Z])?$")


@dataclass
class WorkbookScan:
    filename: str
    file_type: str
    sheets: list[dict[str, Any]]
    method_counts: Counter[str]
    formula_functions: Counter[str]
    private_field_count: int
    ticker_like_count: int


def _norm(value: Any) -> str:
    return str(value or "").strip().lower()


def _classify_text(value: Any) -> set[str]:
    text = _norm(value)
    if not text:
        return set()
    hits = set()
    for category, terms in METHODOLOGY_PATTERNS.items():
        if any(term in text for term in terms):
            hits.add(category)
    return hits


def _is_private_header(value: Any) -> bool:
    text = _norm(value)
    return bool(text and any(term in text for term in PRIVATE_FIELD_PATTERNS))


def _formula_functions(formula: str) -> list[str]:
    return [match.group(1).upper() for match in FORMULA_FUNCTION_RE.finditer(formula)]


def scan_workbook_bytes(filename: str, content: bytes) -> WorkbookScan:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError("Workbook is too large for private in-memory analysis.")
    if suffix in {"xlsx", "xlsm"}:
        return _scan_xlsx(filename, content, suffix)
    if suffix == "csv":
        return _scan_csv(filename, content)
    raise ValueError("Unsupported upload type. Use .xlsx, .xlsm, or .csv.")


def _scan_xlsx(filename: str, content: bytes, suffix: str) -> WorkbookScan:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    sheets = []
    method_counts: Counter[str] = Counter()
    function_counts: Counter[str] = Counter()
    private_fields = 0
    ticker_like = 0

    for idx, sheet in enumerate(workbook.worksheets, start=1):
        nonempty = 0
        formula_count = 0
        sheet_methods: Counter[str] = Counter()
        sheet_private = 0
        header_rows_checked = 0
        max_row = min(sheet.max_row or 0, MAX_SCAN_ROWS)
        max_col = min(sheet.max_column or 0, MAX_SCAN_COLUMNS)

        for row_index, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col),
            start=1,
        ):
            row_values = []
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                nonempty += 1
                row_values.append(value)
                for method in _classify_text(value):
                    method_counts[method] += 1
                    sheet_methods[method] += 1
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    for fn in _formula_functions(value):
                        function_counts[fn] += 1
                elif isinstance(value, str) and TICKER_LIKE_RE.match(value.strip()):
                    ticker_like += 1
            if row_index <= 5 and row_values:
                header_rows_checked += 1
                row_private = sum(1 for value in row_values if _is_private_header(value))
                private_fields += row_private
                sheet_private += row_private

        sheets.append({
            "sheet": f"Sheet {idx}",
            "rows": sheet.max_row or 0,
            "columns": sheet.max_column or 0,
            "sampled_rows": max_row,
            "sampled_columns": max_col,
            "nonempty_cells_sampled": nonempty,
            "formula_cells_sampled": formula_count,
            "header_rows_checked": header_rows_checked,
            "private_columns_redacted": sheet_private,
            "method_signals": sorted(sheet_methods),
        })

    return WorkbookScan(
        filename=filename,
        file_type=suffix,
        sheets=sheets,
        method_counts=method_counts,
        formula_functions=function_counts,
        private_field_count=private_fields,
        ticker_like_count=ticker_like,
    )


def _scan_csv(filename: str, content: bytes) -> WorkbookScan:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(StringIO(text))
    rows = []
    for idx, row in enumerate(reader):
        if idx >= MAX_SCAN_ROWS:
            break
        rows.append(row[:MAX_SCAN_COLUMNS])

    method_counts: Counter[str] = Counter()
    private_fields = 0
    ticker_like = 0
    nonempty = 0
    for row_index, row in enumerate(rows, start=1):
        for value in row:
            if value == "":
                continue
            nonempty += 1
            for method in _classify_text(value):
                method_counts[method] += 1
            if row_index <= 5 and _is_private_header(value):
                private_fields += 1
            if TICKER_LIKE_RE.match(str(value).strip()):
                ticker_like += 1

    sheets = [{
        "sheet": "Sheet 1",
        "rows": len(rows),
        "columns": max((len(row) for row in rows), default=0),
        "sampled_rows": len(rows),
        "sampled_columns": min(max((len(row) for row in rows), default=0), MAX_SCAN_COLUMNS),
        "nonempty_cells_sampled": nonempty,
        "formula_cells_sampled": 0,
        "header_rows_checked": min(len(rows), 5),
        "private_columns_redacted": private_fields,
        "method_signals": sorted(method_counts),
    }]

    return WorkbookScan(
        filename=filename,
        file_type="csv",
        sheets=sheets,
        method_counts=method_counts,
        formula_functions=Counter(),
        private_field_count=private_fields,
        ticker_like_count=ticker_like,
    )


def build_sanitized_master_plan(
    scan: WorkbookScan,
    recommendation: dict[str, Any] | None = None,
) -> dict[str, Any]:
    method_signals = [
        {"id": key, "label": _label_method(key), "strength": count}
        for key, count in scan.method_counts.most_common()
    ]
    formula_functions = [
        {"function": name, "count": count}
        for name, count in scan.formula_functions.most_common(12)
    ]
    privacy = {
        "status": "sanitized",
        "raw_holdings_returned": False,
        "raw_account_values_returned": False,
        "raw_sheet_names_returned": False,
        "private_columns_redacted": scan.private_field_count,
        "ticker_like_values_seen_not_returned": scan.ticker_like_count,
        "policy": (
            "Workbook rows, account sizes, share counts, cost basis, broker fields, "
            "notes, and raw tickers from the upload are not returned by this endpoint."
        ),
    }

    plan_steps = [
        {
            "step": "Translate the workbook into Dad Method signals",
            "action": "Map detected chart, rank, valuation, risk, and allocation columns into explicit scoring rules.",
        },
        {
            "step": "Keep holdings private",
            "action": "Use only generalized rules and $1M-normalized examples in GRID; never publish raw workbook rows.",
        },
        {
            "step": "Score two universes weekly",
            "action": "Run the core compounder list and the frontier infrastructure list as separate boards.",
        },
        {
            "step": "Build a $1M model allocation",
            "action": "Apply max-position, cash, hold-buffer, and drawdown guardrails before producing any allocation.",
        },
        {
            "step": "Export a review packet",
            "action": "Generate Excel/PDF packets with chart scores, changes, and a sanitized decision log.",
        },
        {
            "step": "Add broker import only after OAuth is explicit",
            "action": "Treat thinkorswim as Schwab-era OAuth/API work; import account data only into private local storage.",
        },
    ]

    candidate_boards = []
    if recommendation:
        for board in recommendation.get("candidate_boards", []):
            candidate_boards.append({
                "id": board.get("id"),
                "label": board.get("label"),
                "ranked_candidates": board.get("universe", {}).get("ranked_candidates", 0),
                "top_candidates": [
                    {
                        "ticker": row.get("ticker"),
                        "score": row.get("score"),
                        "themes": row.get("themes", []),
                        "years": row.get("years"),
                    }
                    for row in board.get("ranked", [])[:12]
                ],
            })

    return {
        "status": "ok",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "filename": scan.filename,
        "file_type": scan.file_type,
        "privacy": privacy,
        "workbook_summary": {
            "sheet_count": len(scan.sheets),
            "sheets": scan.sheets,
            "method_signals": method_signals,
            "formula_functions": formula_functions,
        },
        "master_plan": {
            "title": "Dad Method Master Plan",
            "objective": (
                "Convert the uploaded workbook into a private methodology layer, "
                "then publish only sanitized rules and a $1M model portfolio."
            ),
            "steps": plan_steps,
            "broker_integration": {
                "label": "thinkorswim / Schwab",
                "status": "credential-gated",
                "next_action": "Confirm Schwab Developer API credentials and OAuth callback before importing account data.",
            },
        },
        "candidate_boards": candidate_boards,
    }


def _label_method(key: str) -> str:
    return key.replace("_", " ").title()


def build_plan_export_workbook(
    plan: dict[str, Any],
    recommendation: dict[str, Any] | None = None,
) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Master Plan"
    _append_rows(summary, [
        ("Dad Method Master Plan", ""),
        ("Generated", plan.get("generated_at", "")),
        ("File Type", plan.get("file_type", "")),
        ("Privacy Status", plan.get("privacy", {}).get("status", "")),
        ("Raw Holdings Returned", "No"),
        ("Objective", plan.get("master_plan", {}).get("objective", "")),
    ])
    summary.append(())
    summary.append(("Step", "Action"))
    for item in plan.get("master_plan", {}).get("steps", []):
        summary.append((item.get("step", ""), item.get("action", "")))

    methods = workbook.create_sheet("Method Signals")
    methods.append(("Signal", "Strength"))
    for item in plan.get("workbook_summary", {}).get("method_signals", []):
        methods.append((item.get("label", ""), item.get("strength", 0)))
    methods.append(())
    methods.append(("Formula Function", "Count"))
    for item in plan.get("workbook_summary", {}).get("formula_functions", []):
        methods.append((item.get("function", ""), item.get("count", 0)))

    boards = workbook.create_sheet("Candidate Boards")
    boards.append(("Board", "Rank", "Ticker", "Score", "Themes", "Years"))
    for board in plan.get("candidate_boards", []):
        for idx, row in enumerate(board.get("top_candidates", []), start=1):
            boards.append((
                board.get("label", ""),
                idx,
                row.get("ticker", ""),
                row.get("score", ""),
                ", ".join(row.get("themes", [])),
                row.get("years", ""),
            ))

    allocation = workbook.create_sheet("Model Allocation")
    allocation.append(("Profile", "Rank", "Ticker", "Target Dollars", "Weight", "Shares", "Score"))
    if recommendation:
        for profile in recommendation.get("profiles", []):
            for idx, row in enumerate(profile.get("allocations", []), start=1):
                allocation.append((
                    profile.get("label", ""),
                    idx,
                    row.get("ticker", ""),
                    row.get("target_dollars", ""),
                    row.get("target_weight", ""),
                    row.get("whole_shares", ""),
                    row.get("score", ""),
                ))

    privacy = workbook.create_sheet("Privacy Notes")
    for key, value in plan.get("privacy", {}).items():
        privacy.append((key, str(value)))

    for sheet in workbook.worksheets:
        _style_sheet(sheet)
        _autosize(sheet)

    out = BytesIO()
    workbook.save(out)
    return out.getvalue()


def _append_rows(sheet, rows: list[tuple[Any, ...]]) -> None:
    for row in rows:
        sheet.append(row)


def _style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="102027")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    sheet.freeze_panes = "A2"


def _autosize(sheet) -> None:
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(
            max((len(str(cell.value)) if cell.value is not None else 0 for cell in column), default=0) + 2,
            60,
        )
        sheet.column_dimensions[letter].width = max(width, 12)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
