from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook, load_workbook

from strategy.portfolio_workbook_plan import (
    build_plan_export_workbook,
    build_sanitized_master_plan,
    scan_workbook_bytes,
)


def _sample_workbook_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Private Holdings"
    sheet.append(["Ticker", "Shares", "Cost Basis", "10 Year Trend", "QQQ Relative", "Rank"])
    sheet.append(["AAPL", 100, 15000, 0.22, 0.05, "=RANK(D2,D:D)"])
    sheet.append(["NVDA", 20, 18000, 0.55, 0.21, "=RANK(D3,D:D)"])
    out = BytesIO()
    workbook.save(out)
    return out.getvalue()


def test_workbook_scan_redacts_private_holdings_and_detects_methodology():
    scan = scan_workbook_bytes("dad.xlsx", _sample_workbook_bytes())
    plan = build_sanitized_master_plan(scan)

    assert plan["privacy"]["raw_holdings_returned"] is False
    assert plan["privacy"]["private_columns_redacted"] >= 3
    assert plan["privacy"]["ticker_like_values_seen_not_returned"] >= 2
    signal_ids = {item["id"] for item in plan["workbook_summary"]["method_signals"]}
    assert "long_term_chart" in signal_ids
    assert "relative_strength" in signal_ids
    assert "ranking" in signal_ids
    assert "AAPL" not in str(plan)
    assert "NVDA" not in str(plan)


def test_sanitized_plan_export_is_valid_xlsx():
    scan = scan_workbook_bytes("dad.xlsx", _sample_workbook_bytes())
    plan = build_sanitized_master_plan(scan)

    exported = build_plan_export_workbook(plan)
    workbook = load_workbook(BytesIO(exported), read_only=True)

    assert "Master Plan" in workbook.sheetnames
    assert "Privacy Notes" in workbook.sheetnames
